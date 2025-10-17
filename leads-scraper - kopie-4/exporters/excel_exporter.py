"""
Excel Exporter - POLISHED VERSION
Clean, organized output with niche filtering
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Export leads to clean, filterable Excel"""

    def __init__(self, output_dir: str = "data/exports", logger: logging.Logger = None):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logger or logging.getLogger(__name__)

    def export(self, businesses: List[Dict], filename: str = None, niche: str = "leads",
               location: str = "CZ") -> str:
        """
        Export businesses to polished Excel
        
        Args:
            businesses: List of business dictionaries
            filename: Custom filename
            niche: Business niche
            location: Location
            
        Returns:
            Path to exported file
        """
        if not businesses:
            self.logger.warning("⚠️  No businesses to export")
            return ""

        # Generate filename
        if not filename:
            date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
            filename = f"{location}_{niche}_{date_str}.xlsx"

        output_path = self.output_dir / filename

        self.logger.info(f"📊 Exporting {len(businesses)} businesses...")

        try:
            # Create DataFrame
            df = pd.DataFrame(businesses)
            
            # CLEAN COLUMNS - Remove city and notes
            export_columns = [
                'business_name',
                'niche',
                'phone',
                'email',
                'website',
                'instagram',
                'facebook',
                'address',
                'google_rating',
                'priority_score',
            ]
            
            # Ensure columns exist
            for col in export_columns:
                if col not in df.columns:
                    df[col] = ''
            
            # Filter to only export columns
            df = df[export_columns]
            
            # Rename for clarity
            df.columns = [
                'Business Name',
                'Type',
                'Phone',
                'Email',
                'Website',
                'Instagram',
                'Facebook',
                'Address',
                'Rating',
                'Priority',
            ]
            
            # Categorize
            no_website = df[df['Website'].isna() | (df['Website'] == '')]
            has_website = df[df['Website'].notna() & (df['Website'] != '')]
            high_priority = df[df['Priority'] >= 75]
            
            # Create Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # SHEET 1: ALL LEADS (Main filterable sheet)
                df.to_excel(writer, sheet_name='📊 ALL LEADS', index=False)
                
                # SHEET 2: NO WEBSITE - Priority
                if not no_website.empty:
                    no_website.to_excel(writer, sheet_name='🔥 NO WEBSITE', index=False)
                
                # SHEET 3: HAS WEBSITE
                if not has_website.empty:
                    has_website.to_excel(writer, sheet_name='🌐 HAS WEBSITE', index=False)
                
                # SHEET 4: High Priority
                if not high_priority.empty:
                    high_priority.to_excel(writer, sheet_name='⭐ HIGH PRIORITY', index=False)
                
                # SHEET 5: Summary
                summary = self._create_summary(businesses)
                summary.to_excel(writer, sheet_name='📈 SUMMARY', index=False)

            # Apply formatting
            self._apply_formatting(output_path)

            self.logger.info(f"✓ Export complete: {output_path}")
            
            # Print summary
            print(f"\n{'='*60}")
            print(f"📊 EXPORT COMPLETE")
            print(f"{'='*60}")
            print(f"🔥 NO WEBSITE: {len(no_website)}")
            print(f"🌐 HAS WEBSITE: {len(has_website)}")
            print(f"📊 TOTAL: {len(businesses)}")
            print(f"📁 FILE: {output_path}")
            print(f"{'='*60}\n")
            print(f"💡 TIP: Use '📊 ALL LEADS' sheet and filter by 'Type' column")
            print(f"   to organize by niche!\n")

            return str(output_path)

        except Exception as e:
            self.logger.error(f"❌ Export error: {e}")
            
            # Emergency CSV backup
            try:
                backup_path = self.output_dir / f"BACKUP_{filename.replace('.xlsx', '.csv')}"
                df.to_csv(backup_path, index=False, encoding='utf-8')
                self.logger.info(f"✅ Backup saved: {backup_path}")
                return str(backup_path)
            except:
                raise

    def _create_summary(self, businesses: List[Dict]) -> pd.DataFrame:
        """Create summary statistics"""
        total = len(businesses)
        
        # By website status
        no_website = sum(1 for b in businesses if not b.get('website'))
        has_website = total - no_website
        
        # By priority
        high_priority = sum(1 for b in businesses if b.get('priority_score', 0) >= 75)
        
        # By contact info
        with_phone = sum(1 for b in businesses if b.get('phone'))
        with_email = sum(1 for b in businesses if b.get('email'))
        with_instagram = sum(1 for b in businesses if b.get('instagram'))
        with_facebook = sum(1 for b in businesses if b.get('facebook'))
        
        # By niche
        niches = {}
        for b in businesses:
            niche = b.get('niche', 'Unknown')
            niches[niche] = niches.get(niche, 0) + 1
        
        # Average rating
        ratings = [b.get('google_rating') for b in businesses if b.get('google_rating')]
        avg_rating = sum(ratings) / len(ratings) if ratings else 0

        summary_data = {
            'Metric': [
                '📊 TOTAL BUSINESSES',
                '',
                '🔥 No Website (QUALIFIED)',
                '🌐 Has Website',
                '⭐ High Priority (75+)',
                '',
                '📞 CONTACT INFO',
                '',
                'With Phone',
                'With Email',
                'With Instagram',
                'With Facebook',
                '',
                '⭐ GOOGLE RATING',
                '',
                'Average Rating',
                'Businesses Rated',
                '',
                '🎯 BY NICHE',
                '',
            ] + [f"  {niche}" for niche in sorted(niches.keys())],
            
            'Count': [
                total,
                '',
                no_website,
                has_website,
                high_priority,
                '',
                '',
                '',
                with_phone,
                with_email,
                with_instagram,
                with_facebook,
                '',
                '',
                '',
                f'{avg_rating:.2f} ⭐' if avg_rating else 'N/A',
                len(ratings),
                '',
                '',
                '',
            ] + [niches[niche] for niche in sorted(niches.keys())],
            
            'Percentage': [
                '100%',
                '',
                f'{no_website/total*100:.1f}%' if total else '0%',
                f'{has_website/total*100:.1f}%' if total else '0%',
                f'{high_priority/total*100:.1f}%' if total else '0%',
                '',
                '',
                '',
                f'{with_phone/total*100:.1f}%' if total else '0%',
                f'{with_email/total*100:.1f}%' if total else '0%',
                f'{with_instagram/total*100:.1f}%' if total else '0%',
                f'{with_facebook/total*100:.1f}%' if total else '0%',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
            ] + [f'{niches[niche]/total*100:.1f}%' if total else '0%' for niche in sorted(niches.keys())],
        }

        return pd.DataFrame(summary_data)

    def _apply_formatting(self, filepath: Path) -> None:
        """Apply professional formatting"""
        try:
            workbook = load_workbook(filepath)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Header colors
                if 'NO WEBSITE' in sheet_name:
                    header_color = "FF6B6B"
                elif 'HAS WEBSITE' in sheet_name:
                    header_color = "4A90E2"
                elif 'HIGH PRIORITY' in sheet_name:
                    header_color = "E67E22"
                elif 'SUMMARY' in sheet_name:
                    header_color = "27AE60"
                else:
                    header_color = "5B9BD5"

                header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)

                # Format headers
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    adjusted_width = min(max_length + 3, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Freeze top row
                sheet.freeze_panes = 'A2'

                # Add filters
                if sheet.max_row > 1:
                    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

            workbook.save(filepath)

        except Exception as e:
            self.logger.warning(f"⚠️  Formatting failed: {e}")