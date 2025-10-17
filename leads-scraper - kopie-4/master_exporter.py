#!/usr/bin/env python3
"""
Master Exporter - Export ALL database leads to ONE Excel file
With filtering by city, niche, priority
"""
import sys
from pathlib import Path
from datetime import datetime
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent))

from colorama import Fore, Style, init
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database_manager import DatabaseManager

init(autoreset=True)


class MasterExporter:
    """Export all scraped leads to ONE comprehensive Excel file"""
    
    def __init__(self, db: DatabaseManager):
        self.db = db
        self.output_dir = Path("data/exports")
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_master_file(self, niche: str = None, city: str = None, 
                          filename: str = None) -> str:
        """
        Export ALL leads from database to ONE master Excel file
        
        Args:
            niche: Optional filter by niche
            city: Optional filter by city
            filename: Custom filename (optional)
            
        Returns:
            Path to exported file
        """
        print(f"\n{Fore.CYAN}📊 EXPORTING MASTER DATABASE{Style.RESET_ALL}")
        print(f"{Fore.CYAN}{'='*70}{Style.RESET_ALL}")
        
        # Get all businesses from database
        businesses = self.db.get_businesses(niche=niche, city=city)
        
        if not businesses:
            print(f"{Fore.RED}❌ No businesses found in database{Style.RESET_ALL}")
            return ""
        
        print(f"✓ Found {Fore.GREEN}{len(businesses)}{Style.RESET_ALL} businesses in database")
        
        # Generate filename
        if not filename:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
            niche_str = niche or "ALL"
            city_str = city or "ALL"
            filename = f"MASTER_{niche_str}_{city_str}_{timestamp}.xlsx"
        
        output_path = self.output_dir / filename
        
        # Convert to DataFrame
        df = pd.DataFrame(businesses)
        
        # Select and order columns
        columns = [
            'business_name',
            'phone',
            'email',
            'website',
            'instagram',
            'facebook',
            'address',
            'city',
            'postal_code',
            'google_rating',
            'niche',
            'source',
            'notes',
            'first_scraped_at',
        ]
        
        # Ensure all columns exist
        for col in columns:
            if col not in df.columns:
                df[col] = ''
        
        df = df[columns]
        
        # Categorize businesses
        no_website = df[df['website'].isna() | (df['website'] == '')]
        has_website = df[df['website'].notna() & (df['website'] != '')]
        
        # Get unique cities and niches for filtering
        cities = df['city'].unique().tolist()
        niches = df['niche'].unique().tolist()
        
        print(f"\n{Fore.YELLOW}📋 Creating Excel sheets:{Style.RESET_ALL}")
        print(f"   • 🔥 NO WEBSITE - QUALIFIED ({len(no_website)} leads)")
        print(f"   • 🌐 HAS WEBSITE ({len(has_website)} leads)")
        print(f"   • 📊 ALL LEADS ({len(df)} total)")
        print(f"   • 🏙️  By City sheets ({len(cities)} cities)")
        print(f"   • 🎯 By Niche sheets ({len(niches)} niches)")
        
        # Create Excel with multiple sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Sheet 1: NO WEBSITE - Qualified Leads
            if not no_website.empty:
                no_website.to_excel(writer, sheet_name='🔥 NO WEBSITE - QUALIFIED', index=False)
            
            # Sheet 2: HAS WEBSITE
            if not has_website.empty:
                has_website.to_excel(writer, sheet_name='🌐 HAS WEBSITE', index=False)
            
            # Sheet 3: ALL LEADS
            df.to_excel(writer, sheet_name='📊 ALL LEADS', index=False)
            
            # Sheets by CITY
            for city_name in sorted(cities):
                if pd.notna(city_name) and city_name:
                    city_df = df[df['city'] == city_name]
                    if not city_df.empty:
                        # Truncate sheet name if too long
                        sheet_name = f"🏙️ {city_name}"[:31]
                        city_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Sheets by NICHE
            for niche_name in sorted(niches):
                if pd.notna(niche_name) and niche_name:
                    niche_df = df[df['niche'] == niche_name]
                    if not niche_df.empty:
                        # Truncate sheet name if too long
                        sheet_name = f"🎯 {niche_name}"[:31]
                        niche_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Summary sheet
            summary_df = self._create_summary(df)
            summary_df.to_excel(writer, sheet_name='📈 SUMMARY', index=False)
        
        # Apply formatting
        print(f"\n{Fore.YELLOW}🎨 Applying formatting...{Style.RESET_ALL}")
        self._apply_formatting(output_path)
        
        # Final summary
        print(f"\n{Fore.GREEN}{'='*70}")
        print(f"✅ MASTER EXPORT COMPLETE!")
        print(f"{'='*70}{Style.RESET_ALL}")
        print(f"📁 File: {Fore.CYAN}{output_path}{Style.RESET_ALL}")
        print(f"📊 Total Leads: {Fore.CYAN}{len(df)}{Style.RESET_ALL}")
        print(f"🔥 No Website: {Fore.GREEN}{len(no_website)}{Style.RESET_ALL}")
        print(f"🌐 Has Website: {Fore.BLUE}{len(has_website)}{Style.RESET_ALL}")
        print(f"🏙️  Cities: {Fore.YELLOW}{len(cities)}{Style.RESET_ALL}")
        print(f"🎯 Niches: {Fore.YELLOW}{len(niches)}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}{'='*70}{Style.RESET_ALL}\n")
        
        print(f"{Fore.CYAN}💡 How to use this file:{Style.RESET_ALL}")
        print(f"   1. Open in Excel/Google Sheets")
        print(f"   2. Use sheet tabs to filter by City or Niche")
        print(f"   3. '🔥 NO WEBSITE' = Your qualified leads")
        print(f"   4. '🌐 HAS WEBSITE' = Send to analyzer")
        print(f"   5. '📊 ALL LEADS' = Complete database\n")
        
        return str(output_path)
    
    def _create_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create summary statistics"""
        
        total = len(df)
        no_website = len(df[df['website'].isna() | (df['website'] == '')])
        has_website = len(df[df['website'].notna() & (df['website'] != '')])
        
        with_phone = len(df[df['phone'].notna() & (df['phone'] != '')])
        with_email = len(df[df['email'].notna() & (df['email'] != '')])
        
        cities_count = df['city'].nunique()
        niches_count = df['niche'].nunique()
        
        # By city breakdown
        city_counts = df['city'].value_counts().head(10)
        
        # By niche breakdown
        niche_counts = df['niche'].value_counts().head(10)
        
        summary_data = {
            'Metric': [
                '🎯 TOTAL DATABASE',
                '',
                'Total Businesses',
                '🔥 No Website (QUALIFIED)',
                '🌐 Has Website',
                '',
                '📞 CONTACT INFO',
                '',
                'With Phone',
                'With Email',
                '',
                '🗺️ COVERAGE',
                '',
                'Total Cities',
                'Total Niches',
                '',
                '🏙️ TOP 5 CITIES',
            ] + [f"  {city}" for city in city_counts.head(5).index] + [
                '',
                '🎯 TOP 5 NICHES',
            ] + [f"  {niche}" for niche in niche_counts.head(5).index],
            
            'Count': [
                '',
                '',
                total,
                no_website,
                has_website,
                '',
                '',
                '',
                with_phone,
                with_email,
                '',
                '',
                '',
                cities_count,
                niches_count,
                '',
                '',
            ] + [count for count in city_counts.head(5).values] + [
                '',
                '',
            ] + [count for count in niche_counts.head(5).values],
            
            'Percentage': [
                '',
                '',
                '100%',
                f'{no_website/total*100:.1f}%' if total else '0%',
                f'{has_website/total*100:.1f}%' if total else '0%',
                '',
                '',
                '',
                f'{with_phone/total*100:.1f}%' if total else '0%',
                f'{with_email/total*100:.1f}%' if total else '0%',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
            ] + [f'{count/total*100:.1f}%' if total else '0%' for count in city_counts.head(5).values] + [
                '',
                '',
            ] + [f'{count/total*100:.1f}%' if total else '0%' for count in niche_counts.head(5).values],
        }
        
        return pd.DataFrame(summary_data)
    
    def _apply_formatting(self, filepath: Path):
        """Apply professional formatting"""
        try:
            workbook = load_workbook(filepath)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Header color based on sheet type
                if 'NO WEBSITE' in sheet_name:
                    header_color = "FF6B6B"  # Red
                elif 'HAS WEBSITE' in sheet_name:
                    header_color = "4A90E2"  # Blue
                elif 'SUMMARY' in sheet_name:
                    header_color = "27AE60"  # Green
                else:
                    header_color = "366092"  # Default blue
                
                header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                # Format headers
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 3, 60)
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Freeze top row
                sheet.freeze_panes = 'A2'
                
                # Add filters
                if sheet.max_row > 1:
                    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            
            workbook.save(filepath)
            
        except Exception as e:
            print(f"{Fore.YELLOW}⚠️  Could not apply formatting: {e}{Style.RESET_ALL}")


def main():
    """Main export function"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Export master database to Excel")
    parser.add_argument('--niche', type=str, help='Filter by niche')
    parser.add_argument('--city', type=str, help='Filter by city')
    parser.add_argument('--output', type=str, help='Custom output filename')
    
    args = parser.parse_args()
    
    print(f"""
{Fore.CYAN}╔══════════════════════════════════════════════════════════╗
║   📊 MASTER DATABASE EXPORTER                            ║
║   Export ALL scraped leads to ONE Excel file             ║
╚══════════════════════════════════════════════════════════╝{Style.RESET_ALL}
""")
    
    # Initialize database
    db = DatabaseManager()
    exporter = MasterExporter(db)
    
    try:
        # Export
        output_file = exporter.export_master_file(
            niche=args.niche,
            city=args.city,
            filename=args.output
        )
        
        if output_file:
            print(f"{Fore.GREEN}✅ Success! Open your file and start selling! 💰{Style.RESET_ALL}\n")
            return 0
        else:
            return 1
    
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())