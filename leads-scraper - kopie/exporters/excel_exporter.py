"""
Excel Exporter
Exports lead data to professionally formatted Excel files
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Exports business leads to Excel with professional formatting"""

    def __init__(self, output_dir: str = "data/exports", logger: logging.Logger = None):
        """
        Initialize Excel exporter

        Args:
            output_dir: Directory for export files
            logger: Logger instance
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logger or logging.getLogger(__name__)

    def export(self, businesses: List[Dict], filename: str = None, niche: str = "leads",
               location: str = "CZ") -> str:
        """
        Export businesses to Excel file with multiple sheets

        Args:
            businesses: List of business dictionaries
            filename: Custom filename (optional)
            niche: Business niche for filename
            location: Location for filename

        Returns:
            Path to exported file
        """
        if not businesses:
            self.logger.warning("No businesses to export")
            return ""

        # Generate filename
        if not filename:
            date_str = datetime.now().strftime("%Y-%m-%d")
            filename = f"{niche}_{location}_{date_str}.xlsx"

        output_path = self.output_dir / filename

        self.logger.info(f"Exporting {len(businesses)} businesses to {output_path}")

        # Create DataFrames for different priority levels
        all_leads = pd.DataFrame(businesses)

        high_priority = [b for b in businesses if b.get('priority_score', 0) >= 75]
        medium_priority = [b for b in businesses if 50 <= b.get('priority_score', 0) < 75]
        low_priority = [b for b in businesses if b.get('priority_score', 0) < 50]

        # Select columns to export
        columns = [
            'business_name',
            'phone',
            'email',
            'website',
            'address',
            'city',
            'instagram',
            'facebook',
            'google_rating',
            'priority_score',
            'priority_category',
            'notes',
        ]

        # Ensure columns exist
        for col in columns:
            if col not in all_leads.columns:
                all_leads[col] = ''

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: High Priority
            if high_priority:
                high_df = pd.DataFrame(high_priority)[columns]
                high_df.to_excel(writer, sheet_name='High Priority', index=False)

            # Sheet 2: Medium Priority
            if medium_priority:
                medium_df = pd.DataFrame(medium_priority)[columns]
                medium_df.to_excel(writer, sheet_name='Medium Priority', index=False)

            # Sheet 3: Low Priority
            if low_priority:
                low_df = pd.DataFrame(low_priority)[columns]
                low_df.to_excel(writer, sheet_name='Low Priority', index=False)

            # Sheet 4: All Leads
            all_leads[columns].to_excel(writer, sheet_name='All Leads', index=False)

            # Sheet 5: Summary Statistics
            summary = self._create_summary_stats(businesses)
            summary.to_excel(writer, sheet_name='Summary', index=False)

        # Apply formatting
        self._apply_formatting(output_path)

        self.logger.info(f"Export complete: {output_path}")
        return str(output_path)

    def _create_summary_stats(self, businesses: List[Dict]) -> pd.DataFrame:
        """
        Create summary statistics DataFrame

        Args:
            businesses: List of business dictionaries

        Returns:
            Summary DataFrame
        """
        total = len(businesses)

        # Count businesses by priority
        high_priority = sum(1 for b in businesses if b.get('priority_score', 0) >= 75)
        medium_priority = sum(1 for b in businesses if 50 <= b.get('priority_score', 0) < 75)
        low_priority = sum(1 for b in businesses if b.get('priority_score', 0) < 50)

        # Count contact methods
        with_phone = sum(1 for b in businesses if b.get('phone'))
        with_email = sum(1 for b in businesses if b.get('email'))
        with_website = sum(1 for b in businesses if b.get('website'))
        with_instagram = sum(1 for b in businesses if b.get('instagram'))
        with_facebook = sum(1 for b in businesses if b.get('facebook'))

        # Count by website status
        no_website = sum(1 for b in businesses if not b.get('website'))
        poor_website = sum(1 for b in businesses if b.get('website') and b.get('website_quality_score', 0) < 50)

        # Average rating
        ratings = [b.get('google_rating') for b in businesses if b.get('google_rating')]
        avg_rating = sum(ratings) / len(ratings) if ratings else 0

        summary_data = {
            'Metric': [
                'Total Businesses',
                '',
                'Priority Distribution',
                '  High Priority (75+)',
                '  Medium Priority (50-74)',
                '  Low Priority (<50)',
                '',
                'Contact Information',
                '  With Phone',
                '  With Email',
                '  With Website',
                '  With Instagram',
                '  With Facebook',
                '',
                'Website Status',
                '  No Website',
                '  Poor Website (<50 quality)',
                '',
                'Google Rating',
                '  Average Rating',
                '  Businesses with Rating',
            ],
            'Count': [
                total,
                '',
                '',
                high_priority,
                medium_priority,
                low_priority,
                '',
                '',
                with_phone,
                with_email,
                with_website,
                with_instagram,
                with_facebook,
                '',
                '',
                no_website,
                poor_website,
                '',
                '',
                f'{avg_rating:.2f}',
                len(ratings),
            ],
            'Percentage': [
                '100%',
                '',
                '',
                f'{high_priority/total*100:.1f}%' if total else '0%',
                f'{medium_priority/total*100:.1f}%' if total else '0%',
                f'{low_priority/total*100:.1f}%' if total else '0%',
                '',
                '',
                f'{with_phone/total*100:.1f}%' if total else '0%',
                f'{with_email/total*100:.1f}%' if total else '0%',
                f'{with_website/total*100:.1f}%' if total else '0%',
                f'{with_instagram/total*100:.1f}%' if total else '0%',
                f'{with_facebook/total*100:.1f}%' if total else '0%',
                '',
                '',
                f'{no_website/total*100:.1f}%' if total else '0%',
                f'{poor_website/total*100:.1f}%' if total else '0%',
                '',
                '',
                '',
                '',
            ]
        }

        return pd.DataFrame(summary_data)

    def _apply_formatting(self, filepath: Path) -> None:
        """
        Apply professional formatting to Excel file

        Args:
            filepath: Path to Excel file
        """
        try:
            workbook = load_workbook(filepath)

            # Format each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Header formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)

                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)  # Cap at 50
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Color-code priority scores
                if 'priority_score' in [cell.value for cell in sheet[1]]:
                    priority_col_idx = None
                    for idx, cell in enumerate(sheet[1], 1):
                        if cell.value == 'priority_score':
                            priority_col_idx = idx
                            break

                    if priority_col_idx:
                        for row in range(2, sheet.max_row + 1):
                            cell = sheet.cell(row=row, column=priority_col_idx)
                            if cell.value:
                                try:
                                    score = int(cell.value)
                                    if score >= 90:
                                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                                    elif score >= 75:
                                        cell.fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
                                    elif score >= 50:
                                        cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                                    else:
                                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                                except (ValueError, TypeError):
                                    pass

                # Freeze top row
                sheet.freeze_panes = 'A2'

                # Add filters
                sheet.auto_filter.ref = sheet.dimensions

            workbook.save(filepath)
            self.logger.debug(f"Applied formatting to {filepath}")

        except Exception as e:
            self.logger.error(f"Error applying formatting: {e}")
