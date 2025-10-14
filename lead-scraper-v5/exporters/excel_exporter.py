"""
Excel Exporter - CRASH-PROOF VERSION
Exports lead data to professionally formatted Excel files with error recovery
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import traceback


class ExcelExporter:
    """Exports business leads to Excel with professional formatting"""

    def __init__(self, output_dir: str = "data/exports", logger: logging.Logger = None):
        """
        Initialize Excel exporter

        Args:
            output_dir: Directory for export files
            logger: Logger instance
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logger or logging.getLogger(__name__)

    def export(self, businesses: List[Dict], filename: str = None, niche: str = "leads",
               location: str = "CZ") -> str:
        """
        Export businesses to Excel file with multiple sheets

        Args:
            businesses: List of business dictionaries
            filename: Custom filename (optional)
            niche: Business niche for filename
            location: Location for filename

        Returns:
            Path to exported file
        """
        if not businesses:
            self.logger.warning("⚠️  No businesses to export")
            return ""

        # Generate filename
        if not filename:
            date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
            filename = f"{niche}_{location}_{date_str}.xlsx"

        output_path = self.output_dir / filename

        self.logger.info(f"📊 Exporting {len(businesses)} businesses to {output_path}")

        try:
            # CRITICAL: Make sure we have data
            if not businesses:
                self.logger.error("❌ No businesses provided to export!")
                return ""

            self.logger.info(f"   ✓ Received {len(businesses)} businesses")

            # Create DataFrames for different categories
            all_leads = pd.DataFrame(businesses)
            
            self.logger.info(f"   ✓ Created DataFrame with {len(all_leads)} rows")

            # 🔥 PRIMARY CATEGORIZATION: Website vs No Website
            no_website_leads = [b for b in businesses if not b.get('website')]
            has_website_leads = [b for b in businesses if b.get('website')]
            
            self.logger.info(f"   ✓ Categorized: {len(no_website_leads)} no website, {len(has_website_leads)} has website")

            # Further categorize by priority
            high_priority = [b for b in businesses if b.get('priority_score', 0) >= 75]
            medium_priority = [b for b in businesses if 50 <= b.get('priority_score', 0) < 75]
            low_priority = [b for b in businesses if b.get('priority_score', 0) < 50]

            # Select columns to export
            columns = [
                'business_name',
                'phone',
                'email',
                'website',
                'instagram',
                'facebook',
                'address',
                'city',
                'google_rating',
                'priority_score',
                'notes',
            ]

            # Ensure columns exist
            for col in columns:
                if col not in all_leads.columns:
                    all_leads[col] = ''

            self.logger.info(f"   ✓ Prepared columns: {', '.join(columns[:5])}...")

            # Create Excel writer with error handling
            try:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    
                    self.logger.info(f"   ✓ Created Excel writer")
                    
                    # 🎯 SHEET 1: NO WEBSITE - QUALIFIED LEADS (Top Priority)
                    if no_website_leads:
                        self.logger.info(f"   → Writing 'NO WEBSITE' sheet ({len(no_website_leads)} leads)...")
                        no_web_df = pd.DataFrame(no_website_leads)[columns]
                        no_web_df.to_excel(writer, sheet_name='🔥 NO WEBSITE - QUALIFIED', index=False)
                        self.logger.info(f"   ✓ Wrote NO WEBSITE sheet")
                    else:
                        self.logger.warning(f"   ⚠️  No 'NO WEBSITE' leads to export")

                    # 🌐 SHEET 2: HAS WEBSITE - NEEDS ANALYSIS
                    if has_website_leads:
                        self.logger.info(f"   → Writing 'HAS WEBSITE' sheet ({len(has_website_leads)} leads)...")
                        has_web_df = pd.DataFrame(has_website_leads)[columns]
                        has_web_df.to_excel(writer, sheet_name='🌐 HAS WEBSITE - ANALYZE', index=False)
                        self.logger.info(f"   ✓ Wrote HAS WEBSITE sheet")
                    else:
                        self.logger.warning(f"   ⚠️  No 'HAS WEBSITE' leads to export")

                    # 📊 SHEET 3: High Priority (all)
                    if high_priority:
                        self.logger.info(f"   → Writing 'High Priority' sheet ({len(high_priority)} leads)...")
                        high_df = pd.DataFrame(high_priority)[columns]
                        high_df.to_excel(writer, sheet_name='High Priority (75+)', index=False)
                        self.logger.info(f"   ✓ Wrote High Priority sheet")

                    # 📈 SHEET 4: Medium Priority
                    if medium_priority:
                        self.logger.info(f"   → Writing 'Medium Priority' sheet ({len(medium_priority)} leads)...")
                        medium_df = pd.DataFrame(medium_priority)[columns]
                        medium_df.to_excel(writer, sheet_name='Medium Priority', index=False)
                        self.logger.info(f"   ✓ Wrote Medium Priority sheet")

                    # 📉 SHEET 5: Low Priority
                    if low_priority:
                        self.logger.info(f"   → Writing 'Low Priority' sheet ({len(low_priority)} leads)...")
                        low_df = pd.DataFrame(low_priority)[columns]
                        low_df.to_excel(writer, sheet_name='Low Priority', index=False)
                        self.logger.info(f"   ✓ Wrote Low Priority sheet")

                    # 📋 SHEET 6: All Leads
                    self.logger.info(f"   → Writing 'All Leads' sheet ({len(all_leads)} leads)...")
                    all_leads[columns].to_excel(writer, sheet_name='All Leads', index=False)
                    self.logger.info(f"   ✓ Wrote All Leads sheet")

                    # 📈 SHEET 7: Summary Statistics
                    self.logger.info(f"   → Writing 'Summary' sheet...")
                    summary = self._create_summary_stats(businesses)
                    summary.to_excel(writer, sheet_name='Summary', index=False)
                    self.logger.info(f"   ✓ Wrote Summary sheet")

                self.logger.info(f"   ✓ Excel file written successfully")

            except Exception as e:
                self.logger.error(f"❌ Error writing Excel file: {e}")
                self.logger.error(traceback.format_exc())
                raise

            # Apply formatting
            try:
                self.logger.info(f"   → Applying formatting...")
                self._apply_formatting(output_path)
                self.logger.info(f"   ✓ Formatting applied")
            except Exception as e:
                self.logger.warning(f"⚠️  Could not apply formatting: {e}")
                # Don't fail if formatting fails - file is still usable

            self.logger.info(f"✅ Export complete: {output_path}")
            
            # Print summary
            print(f"\n{'='*50}")
            print(f"📊 EXPORT SUMMARY")
            print(f"{'='*50}")
            print(f"🔥 NO WEBSITE (Qualified Leads): {len(no_website_leads)}")
            print(f"🌐 HAS WEBSITE (Send to Analyzer): {len(has_website_leads)}")
            print(f"📊 Total Leads: {len(businesses)}")
            print(f"📁 File: {output_path}")
            print(f"{'='*50}\n")

            return str(output_path)

        except Exception as e:
            self.logger.error(f"❌ CRITICAL ERROR in export: {e}")
            self.logger.error(traceback.format_exc())
            
            # Emergency backup: Save raw CSV
            try:
                backup_path = self.output_dir / f"BACKUP_{filename.replace('.xlsx', '.csv')}"
                self.logger.info(f"💾 Attempting emergency backup to CSV: {backup_path}")
                
                df = pd.DataFrame(businesses)
                df.to_csv(backup_path, index=False, encoding='utf-8')
                
                self.logger.info(f"✅ Emergency backup saved: {backup_path}")
                print(f"\n⚠️  Excel export failed, but saved backup CSV: {backup_path}\n")
                
                return str(backup_path)
                
            except Exception as backup_error:
                self.logger.error(f"❌ Even backup failed: {backup_error}")
                raise

    def _create_summary_stats(self, businesses: List[Dict]) -> pd.DataFrame:
        """
        Create summary statistics DataFrame

        Args:
            businesses: List of business dictionaries

        Returns:
            Summary DataFrame
        """
        total = len(businesses)

        # Count by website status
        no_website = sum(1 for b in businesses if not b.get('website'))
        has_website = sum(1 for b in businesses if b.get('website'))

        # Count businesses by priority
        high_priority = sum(1 for b in businesses if b.get('priority_score', 0) >= 75)
        medium_priority = sum(1 for b in businesses if 50 <= b.get('priority_score', 0) < 75)
        low_priority = sum(1 for b in businesses if b.get('priority_score', 0) < 50)

        # Count contact methods
        with_phone = sum(1 for b in businesses if b.get('phone'))
        with_email = sum(1 for b in businesses if b.get('email'))
        with_instagram = sum(1 for b in businesses if b.get('instagram'))
        with_facebook = sum(1 for b in businesses if b.get('facebook'))

        # Average rating
        ratings = [b.get('google_rating') for b in businesses if b.get('google_rating')]
        avg_rating = sum(ratings) / len(ratings) if ratings else 0

        summary_data = {
            'Metric': [
                '🎯 LEAD CATEGORIZATION',
                '',
                '🔥 NO WEBSITE (Qualified Leads)',
                '🌐 HAS WEBSITE (Send to Analyzer)',
                '',
                '📊 TOTAL BUSINESSES',
                '',
                '━━━━━━━━━━━━━━━━━━━━',
                '',
                'Priority Distribution',
                '  High Priority (75+)',
                '  Medium Priority (50-74)',
                '  Low Priority (<50)',
                '',
                'Contact Information Available',
                '  With Phone',
                '  With Email',
                '  With Instagram',
                '  With Facebook',
                '',
                'Google Rating',
                '  Average Rating',
                '  Businesses with Rating',
            ],
            'Count': [
                '',
                '',
                no_website,
                has_website,
                '',
                total,
                '',
                '',
                '',
                '',
                high_priority,
                medium_priority,
                low_priority,
                '',
                '',
                with_phone,
                with_email,
                with_instagram,
                with_facebook,
                '',
                '',
                f'{avg_rating:.2f} ⭐' if avg_rating else 'N/A',
                len(ratings),
            ],
            'Percentage': [
                '',
                '',
                f'{no_website/total*100:.1f}%' if total else '0%',
                f'{has_website/total*100:.1f}%' if total else '0%',
                '',
                '100%',
                '',
                '',
                '',
                '',
                f'{high_priority/total*100:.1f}%' if total else '0%',
                f'{medium_priority/total*100:.1f}%' if total else '0%',
                f'{low_priority/total*100:.1f}%' if total else '0%',
                '',
                '',
                f'{with_phone/total*100:.1f}%' if total else '0%',
                f'{with_email/total*100:.1f}%' if total else '0%',
                f'{with_instagram/total*100:.1f}%' if total else '0%',
                f'{with_facebook/total*100:.1f}%' if total else '0%',
                '',
                '',
                '',
                '',
            ]
        }

        return pd.DataFrame(summary_data)

    def _apply_formatting(self, filepath: Path) -> None:
        """
        Apply professional formatting to Excel file

        Args:
            filepath: Path to Excel file
        """
        try:
            workbook = load_workbook(filepath)

            # Format each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Different header colors for different sheet types
                if 'NO WEBSITE' in sheet_name:
                    # Red/Orange for qualified leads without websites
                    header_color = "FF6B6B"
                elif 'HAS WEBSITE' in sheet_name:
                    # Blue for leads that need website analysis
                    header_color = "4A90E2"
                elif 'High Priority' in sheet_name:
                    header_color = "E67E22"
                elif 'Summary' in sheet_name:
                    header_color = "27AE60"
                else:
                    header_color = "366092"

                header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)

                # Format headers
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    adjusted_width = min(max_length + 3, 60)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Highlight rows without website in green (qualified leads)
                if sheet_name == 'All Leads' or 'Priority' in sheet_name:
                    for row_idx in range(2, sheet.max_row + 1):
                        # Find website column
                        for col_idx, cell in enumerate(sheet[1], 1):
                            if cell.value == 'website':
                                website_cell = sheet.cell(row=row_idx, column=col_idx)
                                # If no website, highlight entire row in light green
                                if not website_cell.value:
                                    for col in range(1, sheet.max_column + 1):
                                        sheet.cell(row=row_idx, column=col).fill = PatternFill(
                                            start_color="D5F4E6", end_color="D5F4E6", fill_type="solid"
                                        )

                # Freeze top row
                sheet.freeze_panes = 'A2'

                # Add filters
                if sheet.max_row > 1:
                    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

            workbook.save(filepath)
            self.logger.debug(f"Applied formatting to {filepath}")

        except Exception as e:
            self.logger.error(f"Error applying formatting: {e}")
            # Don't raise - formatting failure shouldn't kill the export




